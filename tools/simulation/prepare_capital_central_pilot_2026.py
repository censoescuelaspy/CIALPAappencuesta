"""Reconfigure the CIALPA pilot sample for Capital and Central only.

The historical sample is preserved as the audit baseline. Schools from Alto
Parana are replaced with a deterministic stratified sample from the current
Capital/Central frame, while the 55 existing schools in those departments are
retained. Outputs contain school data only and exclude MEC contact fields.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import shutil
import tempfile
import time
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import prepare_mec_roster_2026 as roster


DOMAIN = {"CAPITAL", "CENTRAL"}
DOMAIN_LABEL = "CAPITAL_CENTRAL"
SELECTION_SEED = 20260716
TARGET_SAMPLE = 86
CONFIDENCE_Z = 1.96
EXPECTED_ERROR = 0.10
EXPECTED_P = 0.50
SOURCE_DATE = "2026-07-16"

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_PRIVATE_DIR = ROOT / "tools" / "simulation" / "private-output"
DEFAULT_REPORT = ROOT / "docs" / "INFORME_MUESTRA_PILOTO_CAPITAL_CENTRAL_2026-07-16.md"

EXTRA_HEADERS = [
    "ESTADO_SELECCION",
    "ORIGEN_SELECCION",
    "CODIGO_REEMPLAZADO",
    "NOMBRE_REEMPLAZADO",
    "DEPARTAMENTO_REEMPLAZADO",
    "SEMILLA_REEMPLAZO",
    "MARCO_DOMINIO",
    "ERROR_MUESTRAL_95_PCT",
    "ALUMNOS_POR_AULA_ORIGINAL",
    "AULAS_EST_ORIGINAL",
]

NAVY = "17365D"
BLUE = "1F4E78"
TEAL = "0F6B78"
ORANGE = "E26B0A"
GREEN = "548235"
WHITE = "FFFFFF"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_ORANGE = "FCE4D6"
LIGHT_YELLOW = "FFF2CC"
LIGHT_GRAY = "F2F2F2"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, type=Path, help="Nuevo XLSX MEC 2026")
    parser.add_argument("--pilot", required=True, type=Path, help="Muestra historica de 86 escuelas")
    parser.add_argument("--output-xlsx", type=Path, help="Libro operativo detallado")
    parser.add_argument("--private-dir", type=Path, default=DEFAULT_PRIVATE_DIR)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--seed", type=int, default=SELECTION_SEED)
    return parser.parse_args()


def margin_of_error(sample: int, population: int) -> float:
    if sample <= 0 or population <= 1 or sample > population:
        raise ValueError("Tamano muestral o poblacional invalido.")
    q = 1 - EXPECTED_P
    finite_correction = (population - sample) / (population - 1)
    return CONFIDENCE_Z * math.sqrt(EXPECTED_P * q / sample * finite_correction)


def apportion(total: int, populations: Counter[str]) -> dict[str, int]:
    population_total = sum(populations.values())
    quotas = {key: total * value / population_total for key, value in populations.items()}
    allocation = {key: math.floor(value) for key, value in quotas.items()}
    remainder = total - sum(allocation.values())
    order = sorted(
        populations,
        key=lambda key: (quotas[key] - allocation[key], populations[key], key),
        reverse=True,
    )
    for key in order[:remainder]:
        allocation[key] += 1
    return allocation


def constrained_stratum_targets(
    frame_counts: Counter[tuple[str, str, str]],
    retained_counts: Counter[tuple[str, str, str]],
    department_frame: Counter[str],
    department_targets: dict[str, int],
) -> dict[tuple[str, str, str], int]:
    targets: dict[tuple[str, str, str], int] = {}
    for department in sorted(department_targets):
        keys = sorted(key for key in frame_counts if key[0] == department)
        current = {key: retained_counts[key] for key in keys}
        expected = {
            key: department_targets[department] * frame_counts[key] / department_frame[department]
            for key in keys
        }
        while sum(current.values()) < department_targets[department]:
            key = max(
                keys,
                key=lambda item: (
                    expected[item] - current[item],
                    frame_counts[item] - current[item],
                    item,
                ),
            )
            if current[key] >= frame_counts[key]:
                raise ValueError(f"No quedan escuelas disponibles en el estrato {key}.")
            current[key] += 1
        targets.update(current)
    return targets


def selection_rank(seed: int, key: tuple[str, str, str], code: str) -> str:
    payload = "|".join([str(seed), *key, code]).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def select_replacements(
    frame: list[roster.School],
    retained: list[roster.School],
    historical_codes: set[str],
    stratum_targets: dict[tuple[str, str, str], int],
    seed: int,
) -> list[roster.School]:
    retained_counts = Counter(
        (school.department_key, school.zone, school.enrollment_group) for school in retained
    )
    selected: list[roster.School] = []
    for key in sorted(stratum_targets):
        needed = stratum_targets[key] - retained_counts[key]
        candidates = [
            school
            for school in frame
            if (school.department_key, school.zone, school.enrollment_group) == key
            and school.code not in historical_codes
        ]
        candidates.sort(key=lambda school: selection_rank(seed, key, school.code))
        if len(candidates) < needed:
            raise ValueError(f"Candidatos insuficientes para {key}: {len(candidates)} < {needed}.")
        selected.extend(candidates[:needed])
    return selected


def balance_district_coverage(
    frame: list[roster.School],
    retained: list[roster.School],
    replacements: list[roster.School],
    seed: int,
) -> tuple[list[roster.School], list[dict[str, str]]]:
    """Ensure every district is represented without changing stratum totals."""
    balanced = list(replacements)
    frame_districts = {
        (school.department_key, roster.normalized(school.district)) for school in frame
    }
    swaps: list[dict[str, str]] = []

    while True:
        final = retained + balanced
        selected_codes = {school.code for school in final}
        district_counts = Counter(
            (school.department_key, roster.normalized(school.district)) for school in final
        )
        missing = sorted(frame_districts - set(district_counts))
        if not missing:
            return balanced, swaps

        progress = False
        for district in missing:
            candidates = [
                school
                for school in frame
                if (school.department_key, roster.normalized(school.district)) == district
                and school.code not in selected_codes
            ]
            candidates.sort(
                key=lambda school: selection_rank(
                    seed,
                    (school.department_key, school.zone, school.enrollment_group),
                    school.code,
                )
            )
            for candidate in candidates:
                key = (candidate.department_key, candidate.zone, candidate.enrollment_group)
                removable = [
                    school
                    for school in balanced
                    if (school.department_key, school.zone, school.enrollment_group) == key
                    and district_counts[
                        (school.department_key, roster.normalized(school.district))
                    ] > 1
                ]
                if not removable:
                    continue
                removable.sort(
                    key=lambda school: (
                        district_counts[
                            (school.department_key, roster.normalized(school.district))
                        ],
                        selection_rank(seed, key, school.code),
                    ),
                    reverse=True,
                )
                outgoing = removable[0]
                balanced[balanced.index(outgoing)] = candidate
                swaps.append(
                    {
                        "district_added": candidate.district,
                        "code_added": candidate.code,
                        "district_removed": outgoing.district,
                        "code_removed": outgoing.code,
                        "stratum": " | ".join(key),
                    }
                )
                progress = True
                break
            if progress:
                break
        if not progress:
            missing_labels = ", ".join(f"{item[0]} / {item[1]}" for item in missing)
            raise ValueError(
                "No se pudo completar la cobertura distrital sin alterar los estratos: "
                + missing_labels
            )


def clean_number(value: float | int | None) -> float | int | None:
    if value is None:
        return None
    return int(value) if float(value).is_integer() else value


def sample_row(
    school: roster.School,
    old: dict[str, Any] | None,
    replaced: dict[str, Any] | None,
    order: int,
    department_frame: Counter[str],
    department_sample: Counter[str],
    stratum_frame: Counter[tuple[str, str, str]],
    stratum_sample: Counter[tuple[str, str, str]],
    target_n: int,
    error: float,
    seed: int,
) -> list[Any]:
    key = (school.department_key, school.zone, school.enrollment_group)
    population = stratum_frame[key]
    sample_count = stratum_sample[key]
    retained = old is not None
    old_stratum = roster.text(old.get("ESTRATO")) if retained else ""
    stratum_changed = retained and roster.normalized(old_stratum) != roster.normalized(school.stratum)
    students_per_classroom = 30
    estimated_classrooms = math.ceil(school.enrollment / students_per_classroom) if school.enrollment else 0

    base = [
        order,
        school.department,
        school.stratum,
        department_frame[school.department_key],
        department_sample[school.department_key],
        population,
        sample_count,
        round(sample_count / population, 12),
        round(population / sample_count, 8),
        school.district,
        school.zone,
        school.locality,
        school.enrollment_group,
        clean_number(school.enrollment),
        school.code,
        school.name,
        students_per_classroom,
        estimated_classrooms,
        school.latitude,
        school.longitude,
        roster.verification_label(school.source),
        old.get("Nh_DEP") if retained else "",
        old.get("nh_DEP") if retained else "",
        old.get("Nh_ESTRATO") if retained else "",
        old.get("nh_ESTRATO") if retained else "",
        old.get("PI") if retained else "",
        old.get("FACTOR_EXP") if retained else "",
        old.get("DEPTO") if retained else "",
        old_stratum,
        old.get("MATRICULA") if retained else "",
        old.get("NOMBRE") if retained else "",
        roster.FRAME_VERSION,
        "redominio_capital_central_reemplazo_alto_parana",
        roster.PILOT_SEED,
        target_n,
        TARGET_SAMPLE,
        "SI" if stratum_changed else ("NO" if retained else "NUEVA"),
    ]
    extra = [
        "RETENIDA" if retained else "REEMPLAZO",
        "MUESTRA_HISTORICA" if retained else "REEMPLAZO_ALTO_PARANA",
        replaced.get("_code", "") if replaced else "",
        replaced.get("NOMBRE", "") if replaced else "",
        replaced.get("DEPTO", "") if replaced else "",
        seed if not retained else "",
        DOMAIN_LABEL,
        round(error * 100, 6),
        old.get("ALUMNOS_POR_AULA") if retained else "",
        old.get("AULAS_EST") if retained else "",
    ]
    return base + extra


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_csv(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def style_header(sheet: Any, row: int, color: str = NAVY) -> None:
    for cell in sheet[row]:
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 30


def auto_width(sheet: Any, maximum: int = 42) -> None:
    for column in range(1, sheet.max_column + 1):
        width = max(
            (len(str(sheet.cell(row=row, column=column).value or "")) for row in range(1, min(sheet.max_row, 600) + 1)),
            default=8,
        )
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), maximum)


def add_table(sheet: Any, name: str, style: str) -> None:
    table = Table(displayName=name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def section(sheet: Any, row: int, title: str, color: str = BLUE) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name="Aptos Display", size=12, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 24
    return row + 1


def wait_until_readable(path: Path, expected_sha: str, attempts: int = 30) -> None:
    last_error: Exception | None = None
    for _attempt in range(attempts):
        try:
            if path.is_file() and sha256_file(path) == expected_sha:
                with zipfile.ZipFile(path) as archive:
                    if archive.testzip() is None:
                        return
        except (OSError, zipfile.BadZipFile) as exc:
            last_error = exc
        time.sleep(0.5)
    raise RuntimeError(f"El XLSX sincronizado no quedo legible: {last_error or 'sin detalle'}")


def write_workbook(
    path: Path,
    headers: list[str],
    rows: list[list[Any]],
    replacements: list[dict[str, Any]],
    metrics: dict[str, Any],
) -> None:
    workbook = Workbook()
    sample = workbook.active
    sample.title = "muestra_piloto_def"
    sample.append(headers)
    for row in rows:
        sample.append(row)
    style_header(sample, 1, NAVY)
    sample.freeze_panes = "C2"
    sample.auto_filter.ref = sample.dimensions
    sample.sheet_properties.tabColor = BLUE
    sample.sheet_view.showGridLines = False
    auto_width(sample)
    sample.column_dimensions[get_column_letter(headers.index("NOMBRE") + 1)].width = 44
    sample.column_dimensions[get_column_letter(headers.index("NOMBRE_REEMPLAZADO") + 1)].width = 44
    for name in ("LAT_DEC", "LNG_DEC"):
        column = headers.index(name) + 1
        for row_number in range(2, sample.max_row + 1):
            sample.cell(row=row_number, column=column).number_format = "0.00000000"
    for name in ("PI", "FACTOR_EXP"):
        column = headers.index(name) + 1
        pattern = "0.000000000000" if name == "PI" else "0.00000000"
        for row_number in range(2, sample.max_row + 1):
            sample.cell(row=row_number, column=column).number_format = pattern
    status_column = get_column_letter(headers.index("ESTADO_SELECCION") + 1)
    sample.conditional_formatting.add(
        f"{status_column}2:{status_column}{sample.max_row}",
        CellIsRule(
            operator="equal",
            formula=['"REEMPLAZO"'],
            fill=PatternFill("solid", fgColor=LIGHT_ORANGE),
            font=Font(bold=True, color="9C0006"),
        ),
    )
    # Do not wrap this sheet in an Excel table: Nh_DEP/nh_DEP differ only by case.
    for row_number in range(2, sample.max_row + 1):
        if row_number % 2 == 0:
            for column in range(1, sample.max_column + 1):
                sample.cell(row=row_number, column=column).fill = PatternFill("solid", fgColor="F7FAFC")

    replacement_sheet = workbook.create_sheet("Reemplazos_Alto_Parana")
    replacement_headers = [
        "ORDEN_MUESTRA",
        "CODIGO_SALIENTE",
        "ESCUELA_SALIENTE",
        "DEPARTAMENTO_SALIENTE",
        "CODIGO_ENTRANTE",
        "ESCUELA_ENTRANTE",
        "DEPARTAMENTO_ENTRANTE",
        "DISTRITO_ENTRANTE",
        "LOCALIDAD_ENTRANTE",
        "ESTRATO_ENTRANTE",
        "MATRICULA_ENTRANTE",
        "SEMILLA",
    ]
    replacement_sheet.append(replacement_headers)
    for item in replacements:
        replacement_sheet.append([item[key] for key in replacement_headers])
    style_header(replacement_sheet, 1, TEAL)
    add_table(replacement_sheet, "tblReemplazosAltoParana", "TableStyleMedium4")
    replacement_sheet.freeze_panes = "E2"
    replacement_sheet.sheet_properties.tabColor = TEAL
    replacement_sheet.sheet_view.showGridLines = False
    auto_width(replacement_sheet, 46)
    replacement_sheet.column_dimensions["C"].width = 46
    replacement_sheet.column_dimensions["F"].width = 46

    summary = workbook.create_sheet("Resumen_metodologico")
    summary.sheet_properties.tabColor = ORANGE
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:E1")
    summary["A1"] = "Muestra piloto CIALPA: Capital y Central"
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 34
    summary.merge_cells("A2:E2")
    summary["A2"] = (
        "Se retienen 55 escuelas y se sustituyen las 31 de Alto Parana con una seleccion "
        "estratificada reproducible de Capital y Central."
    )
    summary["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    summary.row_dimensions[2].height = 34
    row_number = section(summary, 4, "Resultado y criterio")
    for column, value in enumerate(["INDICADOR", "VALOR", "CRITERIO", "ESTADO", "FUENTE"], start=1):
        summary.cell(row=row_number, column=column, value=value)
    style_header(summary, row_number, BLUE)
    controls = [
        ("Marco elegible", metrics["frame_total"], "Capital y Central con coordenadas", "VALIDADO", roster.FRAME_VERSION),
        ("Muestra reducida", metrics["retained_total"], "Sin Alto Parana", "INSUFICIENTE", "Muestra historica"),
        ("Error muestra reducida", metrics["retained_error_pct"], "Mayor al limite de 10%", "NO CUMPLE", "Calculo FPC"),
        ("Tamano teorico minimo", metrics["minimum_sample"], "95%, error 10%, p=0,5", "VALIDADO", "Calculo FPC"),
        ("Muestra final", metrics["sample_total"], "Mantiene el tamano operativo historico", "VALIDADO", "Seleccion 2026"),
        ("Error muestra final", metrics["sample_error_pct"], "Menor al limite de 10%", "CUMPLE", "Calculo FPC"),
        ("Escuelas retenidas", metrics["retained_total"], "Capital y Central historicas", "VALIDADO", "Muestra historica"),
        ("Escuelas reemplazadas", metrics["replacement_total"], "Salen solo las de Alto Parana", "VALIDADO", "Seleccion 2026"),
        ("Semilla", metrics["seed"], "Permite reproducir la seleccion", "VALIDADO", "Script"),
        ("Suma ponderadores", metrics["weight_sum"], "Reproduce el marco de 640", "VALIDADO", "Postestratificacion"),
        ("Distritos cubiertos", metrics["districts_covered"], f"De {metrics['districts_frame']} en el marco", "VALIDADO", "Padron MEC"),
        ("Ajustes de cobertura distrital", metrics["district_balancing_swaps"], "Intercambios dentro del mismo estrato", "VALIDADO", "Seleccion 2026"),
        ("Datos personales", "No incluidos", "Sin responsables, telefonos ni correos", "VALIDADO", "Control de privacidad"),
    ]
    for control in controls:
        row_number += 1
        for column, value in enumerate(control, start=1):
            summary.cell(row=row_number, column=column, value=value)
        summary.cell(row=row_number, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        summary.cell(row=row_number, column=4).font = Font(
            bold=True,
            color=GREEN if control[3] in {"VALIDADO", "CUMPLE"} else ORANGE,
        )
    row_number += 2
    row_number = section(summary, row_number, "Asignacion por departamento")
    for column, value in enumerate(["DEPARTAMENTO", "MARCO", "RETENIDAS", "REEMPLAZOS", "MUESTRA_FINAL"], start=1):
        summary.cell(row=row_number, column=column, value=value)
    style_header(summary, row_number, BLUE)
    for department in sorted(metrics["department_frame"]):
        row_number += 1
        values = [
            department,
            metrics["department_frame"][department],
            metrics["department_retained"][department],
            metrics["department_replacements"][department],
            metrics["department_sample"][department],
        ]
        for column, value in enumerate(values, start=1):
            summary.cell(row=row_number, column=column, value=value)
    row_number += 2
    row_number = section(summary, row_number, "Asignacion por estrato", TEAL)
    for column, value in enumerate(["DEPARTAMENTO / ESTRATO", "MARCO", "RETENIDAS", "REEMPLAZOS", "MUESTRA_FINAL"], start=1):
        summary.cell(row=row_number, column=column, value=value)
    style_header(summary, row_number, TEAL)
    for item in metrics["strata"]:
        row_number += 1
        values = [
            f"{item['department']} | {item['zone']} | {item['group']}",
            item["frame"],
            item["retained"],
            item["replacement"],
            item["sample"],
        ]
        for column, value in enumerate(values, start=1):
            summary.cell(row=row_number, column=column, value=value)
    row_number += 2
    row_number = section(summary, row_number, "Decision metodologica", ORANGE)
    summary.merge_cells(start_row=row_number, start_column=1, end_row=row_number + 2, end_column=5)
    summary.cell(row=row_number, column=1, value=(
        "La muestra de 55 escuelas no cumple el error maximo de 10%. Se mantienen sus integrantes y se "
        "agregan 31 reemplazos seleccionados sin reposicion dentro de estratos Departamento x Zona x "
        "Grupo de matricula. Las escuelas salientes de Alto Parana dejan de integrar el piloto vigente, "
        "pero sus formularios y evidencias historicas no deben eliminarse."
    ))
    summary.cell(row=row_number, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    summary.cell(row=row_number, column=1).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
    summary.cell(row=row_number, column=1).font = Font(bold=True)
    summary.row_dimensions[row_number].height = 58
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 26
    summary.column_dimensions["C"].width = 52
    summary.column_dimensions["D"].width = 22
    summary.column_dimensions["E"].width = 28
    summary.freeze_panes = "A6"

    dictionary = workbook.create_sheet("Diccionario_campos")
    dictionary.sheet_properties.tabColor = GREEN
    dictionary.append(["CAMPO", "DESCRIPCION", "ORIGEN", "USO"])
    for header in headers:
        if header in roster.PILOT_HEADERS:
            origin = "Padron MEC / muestra historica / recalculo"
            description = "Campo compatible con la hoja muestra_piloto_def vigente."
        elif header.endswith("_ORIGINAL"):
            origin = "Muestra historica"
            description = "Valor anterior preservado para auditoria."
        else:
            origin = "Redominio Capital-Central"
            description = "Campo de trazabilidad de la nueva seleccion."
        dictionary.append([header, description, origin, "Operacion y auditoria"])
    style_header(dictionary, 1, GREEN)
    add_table(dictionary, "tblDiccionarioPilotoCapitalCentral", "TableStyleMedium7")
    dictionary.freeze_panes = "A2"
    dictionary.sheet_view.showGridLines = False
    dictionary.column_dimensions["A"].width = 34
    dictionary.column_dimensions["B"].width = 62
    dictionary.column_dimensions["C"].width = 38
    dictionary.column_dimensions["D"].width = 24

    for sheet in workbook.worksheets:
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.oddFooter.left.text = "CIALPA - Piloto Capital y Central"
        sheet.oddFooter.right.text = "Pagina &P de &N"

    now = datetime.now(ZoneInfo("America/Asuncion")).replace(tzinfo=None)
    workbook.properties.creator = "CIALPA"
    workbook.properties.lastModifiedBy = "CIALPA"
    workbook.properties.created = now
    workbook.properties.modified = now
    workbook.properties.title = "Muestra piloto CIALPA Capital y Central"
    workbook.properties.subject = "Redominio y reemplazo de escuelas de Alto Parana"
    workbook.properties.description = "Muestra de 86 escuelas sin datos personales de contacto."

    local_temp = Path(tempfile.gettempdir()) / f"cialpa-pilot-{os.getpid()}.xlsx"
    workbook.save(local_temp)
    workbook.close()
    with zipfile.ZipFile(local_temp) as archive:
        if archive.testzip() is not None:
            raise ValueError("El XLSX generado esta corrupto.")
    check = load_workbook(local_temp, read_only=True, data_only=True)
    if check["muestra_piloto_def"].max_row != TARGET_SAMPLE + 1:
        raise ValueError("El XLSX no contiene las 86 escuelas esperadas.")
    check.close()

    path.parent.mkdir(parents=True, exist_ok=True)
    local_sha = sha256_file(local_temp)
    staging = path.with_name(path.stem + ".tmp.xlsx")
    if staging.exists():
        staging.unlink()
    shutil.copy2(local_temp, staging)
    wait_until_readable(staging, local_sha)
    os.replace(staging, path)
    wait_until_readable(path, local_sha)
    local_temp.unlink()


def write_report(path: Path, metrics: dict[str, Any], source: Path, pilot: Path) -> None:
    strata_rows = "\n".join(
        f"| {item['department']} | {item['zone']} | {item['group']} | {item['frame']} | "
        f"{item['retained']} | {item['replacement']} | {item['sample']} |"
        for item in metrics["strata"]
    )
    content = f"""# Informe de muestra piloto Capital y Central

Fecha: `{SOURCE_DATE}`
Marco: `{roster.FRAME_VERSION}`
Semilla de reemplazo: `{metrics['seed']}`

## Decision

- El piloto queda limitado a **Capital y Central**.
- Al retirar Alto Parana quedan **{metrics['retained_total']} escuelas**, con un error estimado de **{metrics['retained_error_pct']:.2f}%** bajo el criterio 95%, p=0,5 y correccion por poblacion finita.
- El minimo teorico para N={metrics['frame_total']} es **n={metrics['minimum_sample']}**; por tanto, 55 no cumple el limite de 10%.
- Se conservan esas 55 escuelas y se seleccionan **{metrics['replacement_total']} reemplazos**: 5 de Capital y 26 de Central.
- La muestra final mantiene **n={metrics['sample_total']}** y un error estimado de **{metrics['sample_error_pct']:.2f}%**.
- La cobertura territorial alcanza **{metrics['districts_covered']} de {metrics['districts_frame']} distritos**, mediante {metrics['district_balancing_swaps']} intercambios dentro del mismo estrato.
- Las 31 escuelas de Alto Parana dejan de marcarse como piloto vigente, pero sus formularios, evidencias y auditoria deben conservarse como historicos.

## Distribucion

| Departamento | Zona | Grupo de matricula | Marco | Retenidas | Reemplazos | Muestra final |
| --- | --- | --- | ---: | ---: | ---: | ---: |
{strata_rows}

## Metodo

La asignacion departamental es proporcional al marco elegible: Capital 15 y Central 71. Dentro de cada departamento se completa la muestra en los estratos `Zona x Grupo de matricula`, respetando las escuelas ya seleccionadas. Los reemplazos se eligen sin reposicion mediante ranking SHA-256 con semilla fija `{metrics['seed']}`. Como control territorial, se exige al menos una escuela por distrito y cualquier ajuste se realiza dentro del mismo estrato. Esto hace que la seleccion sea reproducible sin publicar contactos ni otros datos personales.

## Fuentes

- Padron MEC: `{source.name}`.
- Muestra historica: `{pilot.name}`.
- Huella de codigos de la muestra final: `{metrics['selection_digest']}`.
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def main() -> None:
    args = parse_args()
    source_path = args.source.resolve()
    pilot_path = args.pilot.resolve()
    if not source_path.is_file():
        raise SystemExit(f"No existe la fuente: {source_path}")
    if not pilot_path.is_file():
        raise SystemExit(f"No existe la muestra historica: {pilot_path}")

    private_dir = args.private_dir.resolve()
    private_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = (
        args.output_xlsx.resolve()
        if args.output_xlsx
        else private_dir / "Muestra_CIALPA_Capital_Central_RUE_2026_2026-07-16.xlsx"
    )
    payload_path = private_dir / "muestra_piloto_capital_central_sheet_payload.json"
    audit_path = private_dir / "muestra_piloto_capital_central_audit.json"
    csv_path = private_dir / "muestra_piloto_capital_central.csv"

    source, _notes = roster.read_source(source_path)
    historical = roster.read_pilot(pilot_path)
    by_code = {school.code: school for school in source}
    missing = [row["_code"] for row in historical if row["_code"] not in by_code]
    if missing:
        raise ValueError(f"Codigos historicos ausentes del marco: {', '.join(missing)}")

    frame = [
        school
        for school in source
        if school.department_key in DOMAIN and school.has_coordinates
    ]
    retained_rows = [row for row in historical if by_code[row["_code"]].department_key in DOMAIN]
    removed_rows = [row for row in historical if by_code[row["_code"]].department_key not in DOMAIN]
    retained = [by_code[row["_code"]] for row in retained_rows]

    if len(frame) != 640 or len(retained_rows) != 55 or len(removed_rows) != 31:
        raise ValueError(
            f"Controles inesperados: marco={len(frame)}, retenidas={len(retained_rows)}, salientes={len(removed_rows)}."
        )

    department_frame = Counter(school.department_key for school in frame)
    department_retained = Counter(school.department_key for school in retained)
    department_targets = apportion(TARGET_SAMPLE, department_frame)
    stratum_frame = Counter(
        (school.department_key, school.zone, school.enrollment_group) for school in frame
    )
    stratum_retained = Counter(
        (school.department_key, school.zone, school.enrollment_group) for school in retained
    )
    stratum_targets = constrained_stratum_targets(
        stratum_frame,
        stratum_retained,
        department_frame,
        department_targets,
    )
    replacements = select_replacements(
        frame,
        retained,
        {row["_code"] for row in historical},
        stratum_targets,
        args.seed,
    )
    replacements, district_swaps = balance_district_coverage(
        frame,
        retained,
        replacements,
        args.seed,
    )
    if len(replacements) != len(removed_rows):
        raise ValueError(f"Se esperaban 31 reemplazos y se obtuvieron {len(replacements)}.")

    replacement_order = sorted(
        replacements,
        key=lambda school: (
            school.department_key,
            school.zone,
            school.enrollment_group,
            selection_rank(
                args.seed,
                (school.department_key, school.zone, school.enrollment_group),
                school.code,
            ),
        ),
    )
    removed_order = sorted(removed_rows, key=lambda row: int(roster.as_number(row["ENUMERA"])))
    replacement_for_order = {
        int(roster.as_number(old["ENUMERA"])): (school, old)
        for old, school in zip(removed_order, replacement_order)
    }
    historical_by_code = {row["_code"]: row for row in retained_rows}

    final_entries: list[tuple[int, roster.School, dict[str, Any] | None, dict[str, Any] | None]] = []
    for old in sorted(historical, key=lambda row: int(roster.as_number(row["ENUMERA"]))):
        order = int(roster.as_number(old["ENUMERA"]))
        school = by_code[old["_code"]]
        if school.department_key in DOMAIN:
            final_entries.append((order, school, old, None))
        else:
            replacement, replaced = replacement_for_order[order]
            final_entries.append((order, replacement, None, replaced))

    final_schools = [entry[1] for entry in final_entries]
    final_codes = [school.code for school in final_schools]
    if len(final_codes) != TARGET_SAMPLE or len(set(final_codes)) != TARGET_SAMPLE:
        raise ValueError("La muestra final no contiene 86 codigos unicos.")
    if any(school.department_key not in DOMAIN or not school.has_coordinates for school in final_schools):
        raise ValueError("La muestra final contiene una escuela fuera del dominio o sin coordenadas.")
    if not {school.code for school in retained}.issubset(set(final_codes)):
        raise ValueError("No se conservaron las 55 escuelas requeridas.")
    if {row["_code"] for row in removed_rows} & set(final_codes):
        raise ValueError("Quedaron escuelas de Alto Parana en la muestra final.")

    department_sample = Counter(school.department_key for school in final_schools)
    stratum_sample = Counter(
        (school.department_key, school.zone, school.enrollment_group) for school in final_schools
    )
    if dict(department_sample) != department_targets or dict(stratum_sample) != stratum_targets:
        raise ValueError("La muestra final no coincide con la asignacion calculada.")

    minimum_sample = roster.sample_size(len(frame))
    retained_error = margin_of_error(len(retained), len(frame))
    sample_error = margin_of_error(len(final_schools), len(frame))
    if retained_error <= EXPECTED_ERROR or sample_error > EXPECTED_ERROR:
        raise ValueError("La conclusion de representatividad no coincide con el criterio configurado.")

    rows = [
        sample_row(
            school,
            old,
            replaced,
            order,
            department_frame,
            department_sample,
            stratum_frame,
            stratum_sample,
            minimum_sample,
            sample_error,
            args.seed,
        )
        for order, school, old, replaced in final_entries
    ]
    headers = roster.PILOT_HEADERS + EXTRA_HEADERS
    weight_sum = sum(float(row[roster.PILOT_HEADERS.index("FACTOR_EXP")]) for row in rows)
    if abs(weight_sum - len(frame)) > 1e-5:
        raise ValueError(f"Los ponderadores no reproducen el marco: {weight_sum} != {len(frame)}")

    replacement_records: list[dict[str, Any]] = []
    for old, school in zip(removed_order, replacement_order):
        replacement_records.append(
            {
                "ORDEN_MUESTRA": int(roster.as_number(old["ENUMERA"])),
                "CODIGO_SALIENTE": old["_code"],
                "ESCUELA_SALIENTE": roster.text(old["NOMBRE"]),
                "DEPARTAMENTO_SALIENTE": roster.text(old["DEPTO"]),
                "CODIGO_ENTRANTE": school.code,
                "ESCUELA_ENTRANTE": school.name,
                "DEPARTAMENTO_ENTRANTE": school.department,
                "DISTRITO_ENTRANTE": school.district,
                "LOCALIDAD_ENTRANTE": school.locality,
                "ESTRATO_ENTRANTE": school.stratum,
                "MATRICULA_ENTRANTE": clean_number(school.enrollment),
                "SEMILLA": args.seed,
            }
        )

    department_replacements = Counter(school.department_key for school in replacements)
    districts_frame = {
        (school.department_key, roster.normalized(school.district)) for school in frame
    }
    districts_sample = {
        (school.department_key, roster.normalized(school.district)) for school in final_schools
    }
    selection_digest = hashlib.sha256("\n".join(final_codes).encode("utf-8")).hexdigest()
    strata_metrics = [
        {
            "department": key[0],
            "zone": key[1],
            "group": key[2],
            "frame": stratum_frame[key],
            "retained": stratum_retained[key],
            "replacement": stratum_sample[key] - stratum_retained[key],
            "sample": stratum_sample[key],
            "weight": round(stratum_frame[key] / stratum_sample[key], 8),
        }
        for key in sorted(stratum_frame)
    ]
    metrics = {
        "generated_at": datetime.now(ZoneInfo("America/Asuncion")).isoformat(timespec="seconds"),
        "frame_version": roster.FRAME_VERSION,
        "domain": sorted(DOMAIN),
        "frame_total": len(frame),
        "retained_total": len(retained),
        "replacement_total": len(replacements),
        "sample_total": len(final_schools),
        "minimum_sample": minimum_sample,
        "retained_error_pct": round(retained_error * 100, 6),
        "sample_error_pct": round(sample_error * 100, 6),
        "seed": args.seed,
        "weight_sum": round(weight_sum, 8),
        "selection_digest": selection_digest,
        "department_frame": dict(sorted(department_frame.items())),
        "department_retained": dict(sorted(department_retained.items())),
        "department_replacements": dict(sorted(department_replacements.items())),
        "department_sample": dict(sorted(department_sample.items())),
        "districts_frame": len(districts_frame),
        "districts_covered": len(districts_sample),
        "district_balancing_swaps": len(district_swaps),
        "district_swaps": district_swaps,
        "strata": strata_metrics,
    }

    payload = {
        "meta": {
            "schema": "cialpa_pilot_capital_central_v1",
            "generated_at": metrics["generated_at"],
            "frame_version": roster.FRAME_VERSION,
            "target_sheet": "muestra_piloto_def",
            "sensitive": False,
            "selection_digest": selection_digest,
        },
        "headers": headers,
        "rows": rows,
    }
    audit = {
        "metrics": metrics,
        "source_sha256": sha256_file(source_path),
        "pilot_sha256": sha256_file(pilot_path),
        "output_xlsx": str(output_xlsx),
        "replacement_records": replacement_records,
    }

    write_workbook(output_xlsx, headers, rows, replacement_records, metrics)
    write_json(payload_path, payload)
    write_json(audit_path, audit)
    write_csv(csv_path, headers, rows)
    write_report(args.report.resolve(), metrics, source_path, pilot_path)

    check = load_workbook(output_xlsx, read_only=True, data_only=True)
    sheet = check["muestra_piloto_def"]
    checked_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    checked_codes = [
        roster.canonical_code(row[checked_headers.index("CODIGO")])
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]
    check.close()
    if checked_codes != final_codes:
        raise ValueError("La relectura del XLSX no coincide con la seleccion calculada.")

    print(
        json.dumps(
            {
                "output_xlsx": str(output_xlsx),
                "payload": str(payload_path),
                "audit": str(audit_path),
                "csv": str(csv_path),
                "report": str(args.report.resolve()),
                "metrics": metrics,
                "xlsx_sha256": sha256_file(output_xlsx),
                "validated_after_save": True,
                "personal_contact_fields": False,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
