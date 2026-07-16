#!/usr/bin/env python3
"""Prepare the MEC 2026 roster, pilot reweighting, and Sheets staging payload.

The source workbook contains personal contact data. Generated private artifacts
are written below tools/simulation/private-output/, which is ignored by Git.
Only the minimal public school index and the aggregate migration report are
written to tracked paths.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("Falta openpyxl. Instale la dependencia en el entorno Python activo.") from exc


APP_VERSION = "2.6.206"
FRAME_VERSION = "RUE_2026_2026-07-16"
SOURCE_DATE = "2026-07-16"
PILOT_SEED = 20260120
PILOT_DOMAIN = {"CAPITAL", "CENTRAL", "ALTO PARANA"}
PILOT_CONFIDENCE_Z = 1.96
PILOT_ERROR = 0.10
PILOT_P = 0.50

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_PUBLIC_JSON = ROOT / "assets" / "data" / "r01-schools-public.json"
DEFAULT_REPORT = ROOT / "docs" / "INFORME_MIGRACION_PADRON_MEC_2026-07-16.md"
DEFAULT_PRIVATE_DIR = ROOT / "tools" / "simulation" / "private-output"

FULL_HEADERS = [
    "orden",
    "codigo_departamento",
    "departamento",
    "codigo_distrito",
    "distrito",
    "localidad",
    "zona",
    "codigo_local",
    "codigo_local_mec",
    "nombre",
    "educacion_inicial",
    "educacion_escolar_basica_1_2_ciclo",
    "educacion_escolar_basica_3_ciclo",
    "educacion_media_bc_bt",
    "educacion_media_abierta",
    "formacion_profesional_media",
    "formacion_vocacional",
    "educacion_permanente_ebbja",
    "educacion_permanente_empja",
    "educacion_permanente_formacion_profesional",
    "servicios_educacion_inclusiva",
    "matricula",
    "responsable_nombre",
    "responsable_contacto",
    "responsable_correo",
    "latitud_dms",
    "longitud_dms",
    "latitud",
    "longitud",
    "marco_version",
    "fuente_archivo",
    "fecha_corte",
]

PILOT_HEADERS = [
    "ENUMERA",
    "DEPTO",
    "ESTRATO",
    "Nh_DEP",
    "nh_DEP",
    "Nh_ESTRATO",
    "nh_ESTRATO",
    "PI",
    "FACTOR_EXP",
    "DIST",
    "ZONA",
    "LOCALIDAD",
    "GRUPO_MATRICULA",
    "MATRICULA",
    "CODIGO",
    "NOMBRE",
    "ALUMNOS_POR_AULA",
    "AULAS_EST",
    "LAT_DEC",
    "LNG_DEC",
    "VERIF",
    "Nh_DEP_ORIGINAL",
    "nh_DEP_ORIGINAL",
    "Nh_ESTRATO_ORIGINAL",
    "nh_ESTRATO_ORIGINAL",
    "PI_ORIGINAL",
    "FACTOR_EXP_ORIGINAL",
    "DEPTO_ORIGINAL",
    "ESTRATO_ORIGINAL",
    "MATRICULA_ORIGINAL",
    "NOMBRE_ORIGINAL",
    "MARCO_VERSION",
    "TIPO_AJUSTE",
    "SEMILLA_ORIGINAL",
    "N_OBJETIVO",
    "N_REALIZADA",
    "CAMBIO_ESTRATO",
]


@dataclass(frozen=True)
class School:
    source: tuple[Any, ...]
    code: str
    display_code: str
    department: str
    department_key: str
    district: str
    locality: str
    zone: str
    name: str
    enrollment: float
    latitude: float | None
    longitude: float | None

    @property
    def has_coordinates(self) -> bool:
        return self.latitude is not None and self.longitude is not None

    @property
    def enrollment_group(self) -> str:
        if self.enrollment < 100:
            return "BAJA(<100)"
        if self.enrollment <= 500:
            return "MEDIA(100-500)"
        return "ALTA(>500)"

    @property
    def stratum(self) -> str:
        return f"{self.zone} , {self.enrollment_group}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, type=Path, help="Nuevo XLSX MEC 2026")
    parser.add_argument("--pilot", required=True, type=Path, help="XLSX de la muestra piloto vigente")
    parser.add_argument(
        "--baseline-public",
        type=Path,
        help="Indice publico anterior usado para calcular altas y bajas; por defecto usa --public-json.",
    )
    parser.add_argument("--public-json", type=Path, default=DEFAULT_PUBLIC_JSON)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--private-dir", type=Path, default=DEFAULT_PRIVATE_DIR)
    return parser.parse_args()


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalized(value: Any) -> str:
    raw = unicodedata.normalize("NFD", text(value).upper())
    return "".join(ch for ch in raw if unicodedata.category(ch) != "Mn").strip()


def canonical_code(value: Any) -> str:
    digits = re.sub(r"\D+", "", text(value))
    if not digits:
        return ""
    return digits.lstrip("0") or "0"


def as_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = text(value).replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def clean_number(value: float) -> int | float:
    return int(value) if float(value).is_integer() else value


def parse_coordinate(value: Any, axis: str) -> float | None:
    if value is None or text(value) == "":
        return None
    if isinstance(value, (int, float)):
        result = float(value)
    else:
        raw = text(value).replace(",", ".")
        dms = re.search(
            r"(-?\d+(?:\.\d+)?)\s*[\u00ba\u00b0]\s*(\d+(?:\.\d+)?)?\s*['\u2032]?\s*(\d+(?:\.\d+)?)?\s*[\"\u2033]?\s*([NSEWO])?",
            raw,
            flags=re.IGNORECASE,
        )
        if dms:
            degrees = abs(float(dms.group(1)))
            minutes = float(dms.group(2) or 0)
            seconds = float(dms.group(3) or 0)
            result = degrees + minutes / 60 + seconds / 3600
            direction = (dms.group(4) or "").upper()
            if dms.group(1).startswith("-") or direction in {"S", "W", "O"}:
                result = -abs(result)
        else:
            try:
                result = float(raw)
            except ValueError:
                return None

    if axis == "lat" and result > 0 and 19 <= result <= 28:
        result = -result
    if axis == "lng" and result > 0 and 54 <= result <= 63:
        result = -result
    if axis == "lat" and not -35 <= result <= 35:
        return None
    if axis == "lng" and not -80 <= result <= 80:
        return None
    return round(result, 8)


def yes(value: Any) -> bool:
    return normalized(value) in {"SI", "S", "YES", "TRUE", "1"} or value == 1


def verification_label(source: tuple[Any, ...]) -> str:
    labels: list[str] = []
    if yes(source[9]):
        labels.append("Inicial")
    if yes(source[10]) or yes(source[11]):
        labels.append("Basica")
    if yes(source[12]) or yes(source[13]):
        labels.append("Media")
    return " ".join(labels) or "Otros"


def read_source(path: Path) -> tuple[list[School], list[list[Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Listado" not in workbook.sheetnames:
        raise ValueError("El XLSX nuevo no contiene la hoja Listado.")
    sheet = workbook["Listado"]
    header = [text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if len(header) < 26 or "codigo" not in normalized(header[7]).lower():
        raise ValueError("La estructura de Listado no coincide con la entrega MEC esperada.")

    rows: list[School] = []
    seen: set[str] = set()
    for row_number, raw in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = tuple(raw[:26])
        if not any(text(value) for value in values):
            continue
        code = canonical_code(values[7])
        if not code:
            raise ValueError(f"Fila {row_number}: registro no vacio sin codigo local.")
        if code in seen:
            raise ValueError(f"Codigo local duplicado en la nueva nomina: {code}.")
        seen.add(code)
        rows.append(
            School(
                source=values,
                code=code,
                display_code=text(values[7]),
                department=text(values[2]),
                department_key=normalized(values[2]),
                district=text(values[4]),
                locality=text(values[5]),
                zone=normalized(values[6]),
                name=text(values[8]),
                enrollment=as_number(values[20]),
                latitude=parse_coordinate(values[24], "lat"),
                longitude=parse_coordinate(values[25], "lng"),
            )
        )

    if len(rows) != 5448:
        raise ValueError(f"Se esperaban 5448 codigos validos y se obtuvieron {len(rows)}.")

    notes: list[list[Any]] = []
    if "notas" in workbook.sheetnames:
        for raw in workbook["notas"].iter_rows(values_only=True):
            row = [value for value in raw]
            if any(text(value) for value in row):
                notes.append(row)
    return rows, notes


def read_pilot(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: idx for idx, header in enumerate(headers)}
    required = set(PILOT_HEADERS[:21])
    missing = sorted(required - set(index))
    if missing:
        raise ValueError(f"Faltan columnas en la muestra piloto: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        code = canonical_code(raw[index["CODIGO"]])
        if not code:
            continue
        rows.append({header: raw[idx] for header, idx in index.items()} | {"_code": code})
    if len(rows) != 86:
        raise ValueError(f"Se esperaban 86 escuelas piloto y se obtuvieron {len(rows)}.")
    return rows


def full_row(school: School, source_name: str) -> list[Any]:
    source = school.source
    return [
        source[0],
        source[1],
        school.department,
        source[3],
        school.district,
        school.locality,
        school.zone,
        school.code,
        school.display_code,
        school.name,
        *source[9:20],
        clean_number(school.enrollment),
        source[21],
        source[22],
        source[23],
        source[24],
        source[25],
        school.latitude,
        school.longitude,
        FRAME_VERSION,
        source_name,
        SOURCE_DATE,
    ]


def build_pilot(
    source: list[School], pilot: list[dict[str, Any]]
) -> tuple[list[list[Any]], dict[str, Any]]:
    by_code = {school.code: school for school in source}
    missing = [row["_code"] for row in pilot if row["_code"] not in by_code]
    if missing:
        raise ValueError(f"Codigos piloto ausentes del marco nuevo: {', '.join(missing)}")

    frame = [
        school
        for school in source
        if school.department_key in PILOT_DOMAIN and school.has_coordinates
    ]
    dep_n = Counter(school.department_key for school in frame)
    stratum_n = Counter((school.department_key, school.zone, school.enrollment_group) for school in frame)

    sample_schools = [by_code[row["_code"]] for row in pilot]
    dep_sample_n = Counter(school.department_key for school in sample_schools)
    stratum_sample_n = Counter(
        (school.department_key, school.zone, school.enrollment_group) for school in sample_schools
    )

    output: list[list[Any]] = []
    changed_strata: list[dict[str, Any]] = []
    name_changes = 0
    enrollment_changes = 0
    locality_changes = 0

    for old, school in zip(pilot, sample_schools):
        key = (school.department_key, school.zone, school.enrollment_group)
        population = stratum_n[key]
        sample_count = stratum_sample_n[key]
        old_stratum = text(old["ESTRATO"])
        stratum_changed = normalized(old_stratum) != normalized(school.stratum)
        if stratum_changed:
            changed_strata.append(
                {
                    "order": int(as_number(old["ENUMERA"])),
                    "code": school.code,
                    "old_stratum": old_stratum,
                    "new_stratum": school.stratum,
                    "old_enrollment": clean_number(as_number(old["MATRICULA"])),
                    "new_enrollment": clean_number(school.enrollment),
                }
            )
        if normalized(old["NOMBRE"]) != normalized(school.name):
            name_changes += 1
        if as_number(old["MATRICULA"]) != school.enrollment:
            enrollment_changes += 1
        if normalized(old["LOCALIDAD"]) != normalized(school.locality):
            locality_changes += 1

        output.append(
            [
                int(as_number(old["ENUMERA"])),
                school.department,
                school.stratum,
                dep_n[school.department_key],
                dep_sample_n[school.department_key],
                population,
                sample_count,
                round(sample_count / population, 12),
                round(population / sample_count, 8),
                school.district,
                school.zone,
                school.locality,
                school.enrollment_group,
                clean_number(school.enrollment),
                school.code,
                school.name,
                old["ALUMNOS_POR_AULA"],
                old["AULAS_EST"],
                school.latitude,
                school.longitude,
                verification_label(school.source),
                old["Nh_DEP"],
                old["nh_DEP"],
                old["Nh_ESTRATO"],
                old["nh_ESTRATO"],
                old["PI"],
                old["FACTOR_EXP"],
                old["DEPTO"],
                old_stratum,
                old["MATRICULA"],
                old["NOMBRE"],
                FRAME_VERSION,
                "postestratificacion_marco_2026_sin_redraw",
                PILOT_SEED,
                sample_size(len(frame)),
                len(pilot),
                "SI" if stratum_changed else "NO",
            ]
        )

    metrics = {
        "frame_total": len(frame),
        "sample_total": len(pilot),
        "sample_target": sample_size(len(frame)),
        "department_frame": dict(sorted(dep_n.items())),
        "department_sample": dict(sorted(dep_sample_n.items())),
        "changed_strata": changed_strata,
        "name_changes": name_changes,
        "enrollment_changes": enrollment_changes,
        "locality_changes": locality_changes,
    }
    return output, metrics


def sample_size(population: int) -> int:
    q = 1 - PILOT_P
    n0 = PILOT_CONFIDENCE_Z**2 * PILOT_P * q / PILOT_ERROR**2
    finite = population * n0 / (population + n0 - 1)
    return math.ceil(finite)


def read_current_public(path: Path) -> dict[str, dict[str, str]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    current: dict[str, dict[str, str]] = {}
    for row in payload.get("schools", []):
        if len(row) < 5:
            continue
        code = canonical_code(row[0])
        if code:
            current[code] = {
                "name": text(row[1]),
                "department": text(row[2]),
                "district": text(row[3]),
                "locality": text(row[4]),
            }
    return current


def public_payload(source: list[School], source_name: str) -> dict[str, Any]:
    schools = sorted(
        (
            [
                school.code,
                school.name,
                school.department,
                school.district,
                school.locality,
                school.code,
            ]
            for school in source
        ),
        key=lambda row: (normalized(row[2]), normalized(row[3]), row[0]),
    )
    return {
        "meta": {
            "schema": "r01_public_school_index_v1",
            "generated_at": SOURCE_DATE,
            "source": source_name,
            "frame_version": FRAME_VERSION,
            "note": "Campos minimos para busqueda publica; no incluye responsables, telefonos ni correos.",
            "total": len(schools),
        },
        "schools": schools,
    }


def aggregate_rows(metrics: dict[str, Any], delta: dict[str, Any], coordinate_count: int) -> list[list[Any]]:
    rows = [
        ["clave", "valor", "observacion"],
        ["app_version", APP_VERSION, "Version preparada para la migracion"],
        ["marco_version", FRAME_VERSION, "Fuente MEC-DGPE RUE 2026"],
        ["total_padron_anterior_app", delta["current_total"], "Indice publico anterior"],
        ["total_padron_nuevo", delta["new_total"], "Codigos locales unicos validos"],
        ["altas", len(delta["added"]), "Codigos que no estaban en el indice publico"],
        ["bajas", len(delta["removed"]), "Codigos que salen del marco vigente"],
        ["con_coordenadas", coordinate_count, "Coordenadas DMS convertidas y validadas"],
        ["sin_coordenadas", delta["new_total"] - coordinate_count, "Sin par completo latitud/longitud"],
        ["muestra_piloto_retenida", metrics["sample_total"], "Los 86 codigos siguen en el marco"],
        ["marco_piloto_2026", metrics["frame_total"], "Dominio Capital, Central y Alto Parana con coordenadas"],
        ["n_objetivo_piloto_2026", metrics["sample_target"], "95%, error 10%, p=0.5, FPC"],
        ["cambios_de_estrato", len(metrics["changed_strata"]), "Por actualizacion de matricula"],
    ]
    for department, population in metrics["department_frame"].items():
        rows.append(
            [
                f"marco_piloto_{department.lower().replace(' ', '_')}",
                population,
                f"Muestra retenida: {metrics['department_sample'].get(department, 0)}",
            ]
        )
    return rows


def build_delta(source: list[School], current: dict[str, dict[str, str]]) -> dict[str, Any]:
    new = {school.code: school for school in source}
    added = [
        {
            "code": code,
            "name": new[code].name,
            "department": new[code].department,
            "district": new[code].district,
        }
        for code in sorted(set(new) - set(current))
    ]
    removed = [
        {"code": code, **current[code]}
        for code in sorted(set(current) - set(new))
    ]
    return {
        "current_total": len(current),
        "new_total": len(new),
        "added": added,
        "removed": removed,
    }


def sheet_payload(
    source_rows: list[list[Any]],
    pilot_rows: list[list[Any]],
    notes: list[list[Any]],
    control_rows: list[list[Any]],
) -> dict[str, Any]:
    stamp = SOURCE_DATE.replace("-", "")
    note_rows = [["Notas de la fuente MEC 2026"]] + notes + [
        [],
        ["Migracion CIALPA", FRAME_VERSION],
        ["Fecha de preparacion", now_iso()],
        ["Criterio piloto", "Se retienen 86 codigos y se recalibran ponderadores; no se redibuja la muestra."],
    ]
    return {
        "meta": {
            "frame_version": FRAME_VERSION,
            "app_version": APP_VERSION,
            "generated_at": now_iso(),
            "sensitive": True,
        },
        "sheets": [
            {
                "title": f"listado_ini_stg_{stamp}",
                "headers": FULL_HEADERS,
                "rows": source_rows,
                "freeze_rows": 1,
                "filter": True,
            },
            {
                "title": f"muestra_piloto_stg_{stamp}",
                "headers": PILOT_HEADERS,
                "rows": pilot_rows,
                "freeze_rows": 1,
                "filter": True,
            },
            {
                "title": f"notas_marco_{stamp}",
                "headers": [],
                "rows": note_rows,
                "freeze_rows": 0,
                "filter": False,
            },
            {
                "title": f"control_migracion_{stamp}",
                "headers": control_rows[0],
                "rows": control_rows[1:],
                "freeze_rows": 1,
                "filter": True,
            },
        ],
    }


def write_private_workbook(path: Path, payload: dict[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for spec in payload["sheets"]:
        sheet = workbook.create_sheet(spec["title"][:31])
        headers = spec["headers"]
        if headers:
            sheet.append(headers)
        for row in spec["rows"]:
            sheet.append(row)
        if headers:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            fill = PatternFill("solid", fgColor="E8EAED")
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = fill
        for column in range(1, min(sheet.max_column, 45) + 1):
            letter = get_column_letter(column)
            longest = max(
                (len(text(sheet.cell(row=row, column=column).value)) for row in range(1, min(sheet.max_row, 250) + 1)),
                default=8,
            )
            sheet.column_dimensions[letter].width = min(max(longest + 2, 10), 42)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_report(
    path: Path,
    source_path: Path,
    pilot_path: Path,
    source: list[School],
    metrics: dict[str, Any],
    delta: dict[str, Any],
) -> None:
    coordinates = sum(school.has_coordinates for school in source)
    changed = "\n".join(
        f"- `{row['code']}`: {row['old_stratum']} -> {row['new_stratum']} "
        f"(matricula {row['old_enrollment']} -> {row['new_enrollment']})."
        for row in metrics["changed_strata"]
    ) or "- Ninguno."
    departments = "\n".join(
        f"| {department} | {population} | {metrics['department_sample'].get(department, 0)} |"
        for department, population in metrics["department_frame"].items()
    )
    content = f"""# Informe de migracion del padron MEC 2026

Fecha de corte: `{SOURCE_DATE}`
Version preparada: `{APP_VERSION}`
Marco: `{FRAME_VERSION}`

## Resultado ejecutivo

- El nuevo archivo contiene **{len(source)} codigos locales validos y unicos**.
- Frente al indice de la app ({delta['current_total']}), incorpora **{len(delta['added'])} altas** y **{len(delta['removed'])} bajas**; el saldo es **{len(source) - delta['current_total']}**.
- Hay **{coordinates}** escuelas con coordenadas validas y **{len(source) - coordinates}** sin un par completo.
- Los **86 codigos** de la muestra piloto permanecen en el nuevo marco.
- Cambiaron **{metrics['enrollment_changes']}** matriculas, **{metrics['name_changes']}** nombres y **{len(metrics['changed_strata'])}** estratos dentro de la muestra.

## Decision sobre la muestra piloto

Se conserva la seleccion historica de 86 escuelas porque todas siguen vigentes y ya existen registros de campo asociados. No se realiza un nuevo sorteo. Los ponderadores se recalibran contra el marco RUE 2026 y los valores originales quedan en columnas separadas.

El marco piloto con coordenadas pasa de `N=972` a `N={metrics['frame_total']}`. Con 95% de confianza, error 10%, `p=0,5` y correccion por poblacion finita, el objetivo sigue siendo `n={metrics['sample_target']}`. La muestra realizada es 86; si el operativo requiere completar el tamano teorico, deben seleccionarse dos reservas adicionales en Central mediante una decision metodologica separada.

| Departamento | Marco 2026 con coordenadas | Muestra retenida |
| --- | ---: | ---: |
{departments}

### Cambios de estrato

{changed}

## Resguardo de datos

- El libro de migracion y el payload de Sheets contienen contactos y quedan en una salida privada ignorada por Git.
- El JSON publico conserva solo codigo, nombre y territorio; no publica responsables, telefonos ni correos.
- Los formularios, cierres, usuarios, auditoria y evidencias no se reemplazan: se vinculan por codigo local.
- Los codigos que salen del marco no deben contar como padron vigente, pero sus registros historicos deben seguir accesibles para administracion.

## Aplicacion del cambio

- La hoja oficial conserva su mismo identificador y ahora expone `listado_ini` con 5448 escuelas y `muestra_piloto_def` con 86 escuelas.
- Las pestañas anteriores se conservaron ocultas como respaldo `legacy_2025_20260716`.
- El backend publico verifico el nuevo marco desde `official_sheet` sin reemplazar las 128 filas operativas existentes.
- GAS version 41 contiene la compatibilidad para consultar registros historicos fuera del padron, pero requiere un deployment desde la cuenta propietaria: los redeploys hechos con la cuenta editora devuelven HTTP 403 pese a declarar acceso anonimo.

## Fuentes

- Nueva nomina MEC: `{source_path.name}`.
- Muestra piloto historica: `{pilot_path.name}`.
- Comparacion de la app: `assets/data/r01-schools-public.json` antes de esta migracion.
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def now_iso() -> str:
    return datetime.now(ZoneInfo("America/Asuncion")).isoformat(timespec="seconds")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_json(path: Path, payload: Any, *, compact: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if compact:
        body = json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n"
    else:
        body = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    path.write_text(body, encoding="utf-8")


def main() -> None:
    args = parse_args()
    source_path = args.source.resolve()
    pilot_path = args.pilot.resolve()
    if not source_path.is_file():
        raise SystemExit(f"No existe la fuente: {source_path}")
    if not pilot_path.is_file():
        raise SystemExit(f"No existe la muestra: {pilot_path}")

    args.private_dir.mkdir(parents=True, exist_ok=True)
    payload_path = args.private_dir / "sheets_migration_payload_2026-07-16.json"
    workbook_path = args.private_dir / "padron_mec_2026_cialpa_migracion_2026-07-16.xlsx"
    audit_path = args.private_dir / "migration_audit_2026-07-16.json"

    source, notes = read_source(source_path)
    pilot = read_pilot(pilot_path)
    baseline_path = (args.baseline_public or args.public_json).resolve()
    baseline_payload = json.loads(baseline_path.read_text(encoding="utf-8"))
    if baseline_payload.get("meta", {}).get("frame_version") == FRAME_VERSION:
        if not audit_path.is_file():
            raise ValueError(
                "El indice publico ya corresponde al marco objetivo y no existe una auditoria previa. "
                "Use --baseline-public con el indice anterior."
            )
        previous_audit = json.loads(audit_path.read_text(encoding="utf-8"))
        delta = previous_audit["delta"]
    else:
        delta = build_delta(source, read_current_public(baseline_path))
    pilot_rows, pilot_metrics = build_pilot(source, pilot)
    source_rows = [full_row(school, source_path.name) for school in source]
    coordinate_count = sum(school.has_coordinates for school in source)
    control = aggregate_rows(pilot_metrics, delta, coordinate_count)
    source_sha256 = sha256_file(source_path)
    pilot_sha256 = sha256_file(pilot_path)
    control.extend([
        ["fuente_sha256", source_sha256, source_path.name],
        ["muestra_sha256", pilot_sha256, pilot_path.name],
    ])
    payload = sheet_payload(source_rows, pilot_rows, notes, control)

    write_json(payload_path, payload)
    write_private_workbook(workbook_path, payload)
    write_json(
        audit_path,
        {
            "generated_at": now_iso(),
            "frame_version": FRAME_VERSION,
            "source_sha256": source_sha256,
            "pilot_sha256": pilot_sha256,
            "totals": {
                "schools": len(source),
                "coordinates": coordinate_count,
                "without_coordinates": len(source) - coordinate_count,
                "added": len(delta["added"]),
                "removed": len(delta["removed"]),
            },
            "delta": delta,
            "pilot": pilot_metrics,
        },
    )
    write_json(args.public_json, public_payload(source, source_path.name), compact=True)
    write_report(args.report, source_path, pilot_path, source, pilot_metrics, delta)

    print(json.dumps({
        "status": "ok",
        "schools": len(source),
        "coordinates": coordinate_count,
        "pilot": len(pilot_rows),
        "pilot_frame": pilot_metrics["frame_total"],
        "pilot_target": pilot_metrics["sample_target"],
        "pilot_stratum_changes": len(pilot_metrics["changed_strata"]),
        "added": len(delta["added"]),
        "removed": len(delta["removed"]),
        "payload": str(payload_path),
        "workbook": str(workbook_path),
        "public_json": str(args.public_json),
        "report": str(args.report),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
